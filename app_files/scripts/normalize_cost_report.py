from __future__ import annotations

from collections import defaultdict
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, Iterable, Optional

from openpyxl import load_workbook


CONTROL_SCOPES = ("rail", "ponds", "roads")
RAIL_PACKAGES = {
    "BOD-BOP-001A",
    "BOD-BOP-001B",
    "BOD-BOP-007D",
    "BOD-BOP-053",
    "BOD-BOP-054",
    "BOD-BOP-061B",
    "BOD-BOP-061F",
    "BOD-BOP-065",
    "BOD-BOP-ROS",
}


def _number(value: Any) -> Optional[float]:
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if value is None:
        return None
    text = str(value).replace("\xa0", " ").strip().replace(" ", "")
    if not text or text.startswith("#"):
        return None
    if text.count(",") == 1 and "." not in text:
        text = text.replace(",", ".")
    else:
        text = text.replace(",", "")
    try:
        return float(text)
    except ValueError:
        return None


def _clean(value: Any) -> str:
    return " ".join(str(value or "").replace("\xa0", " ").split()).strip()


def _headers(row: Iterable[Any]) -> Dict[str, int]:
    return {_clean(value).lower(): index for index, value in enumerate(row) if _clean(value)}


def _column(headers: Dict[str, int], *names: str) -> Optional[int]:
    for name in names:
        if name.lower() in headers:
            return headers[name.lower()]
    return None


def _value(row: list[Any], index: Optional[int]) -> Any:
    return row[index] if index is not None and index < len(row) else None


def _scope_id(main_description: Any, package: Any = None, package_name: Any = None) -> Optional[str]:
    main = _clean(main_description).lower()
    package_text = _clean(package).upper()
    package_name_text = _clean(package_name).lower()
    if main == "ponds and culverts":
        return "ponds"
    if main == "roads and ditches":
        return "roads"
    package_label = package_text.lower()
    if package_text in RAIL_PACKAGES or "rail on site" in package_label or "rails on site" in package_label or "rail on site" in package_name_text or "rails on site" in package_name_text:
        return "rail"
    return None


def _report_date(workbook: Any) -> datetime:
    preferred = ["Budget per Package", "P. Packages", "Consolidated Tracker", "Commitment & budget tracker"]
    for sheet_name in preferred:
        if sheet_name not in workbook.sheetnames:
            continue
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows(min_row=1, max_row=min(8, sheet.max_row), values_only=True):
            dates = [value for value in row if isinstance(value, (datetime, date))]
            if dates:
                return max(datetime.combine(value, datetime.min.time()) if isinstance(value, date) and not isinstance(value, datetime) else value for value in dates)
    raise ValueError("No cost data date was found in the workbook headers.")


def _budget_package_rows(workbook: Any) -> tuple[Dict[str, Dict[str, Any]], list[Dict[str, Any]]]:
    sheet_name = "Budget per Package" if "Budget per Package" in workbook.sheetnames else "P. Packages"
    if sheet_name not in workbook.sheetnames:
        return {}, []
    sheet = workbook[sheet_name]
    header_row = None
    headers: Dict[str, int] = {}
    for row_index, row in enumerate(sheet.iter_rows(min_row=1, max_row=min(12, sheet.max_row), values_only=True), 1):
        candidate = _headers(row)
        if "package number" in candidate and "package description" in candidate:
            header_row, headers = row_index, candidate
            break
    if header_row is None:
        return {}, []

    scope_rows: Dict[str, Dict[str, Any]] = {}
    packages: list[Dict[str, Any]] = []
    package_col = _column(headers, "Package Number")
    description_col = _column(headers, "Package Description")
    supplier_col = _column(headers, "SUPPLIER")
    phase_col = _column(headers, "Phase")
    budget_col = _column(headers, "Approved Control Budget", '"Baseline" Control Budget')
    forecast_col = _column(headers, "EAC - Projected Budget", "EAC")
    exposure_col = _column(headers, "Potential Changes (CRs and Trends)", "Potential Changes")
    potential_col = _column(headers, "FEAC - Potential Forecast")
    vowd_col = _column(headers, "VOWD")

    for raw in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        row = list(raw)
        package = _clean(_value(row, package_col))
        description = _clean(_value(row, description_col))
        if not package and not description:
            continue
        sid = _scope_id(None, package, description)
        if sid != "rail":
            continue
        display_description = description or package
        item = {
            "package": package,
            "description": display_description,
            "supplier": _clean(_value(row, supplier_col)),
            "phase": _clean(_value(row, phase_col)),
            "budget": _number(_value(row, budget_col)),
            "forecast": _number(_value(row, forecast_col)),
            "potential_exposure": _number(_value(row, exposure_col)),
            "potential_forecast": _number(_value(row, potential_col)),
            "vowd": _number(_value(row, vowd_col)),
        }
        item["change_order"] = (item["forecast"] - item["budget"]) if item["forecast"] is not None and item["budget"] is not None else None
        if "total" in display_description.lower():
            scope_rows["rail"] = item
        else:
            packages.append(item)
    return scope_rows, packages


def _tracker_data(workbook: Any) -> tuple[Dict[str, Dict[str, float]], Dict[str, list[Dict[str, Any]]]]:
    totals = {sid: {"committed": 0.0, "vowd": 0.0} for sid in CONTROL_SCOPES}
    packages: Dict[str, Dict[tuple[str, str], Dict[str, Any]]] = {sid: {} for sid in CONTROL_SCOPES}
    if "Commitment & budget tracker" not in workbook.sheetnames:
        return totals, {sid: [] for sid in CONTROL_SCOPES}
    sheet = workbook["Commitment & budget tracker"]
    rows = sheet.iter_rows(values_only=True)
    headers = _headers(next(rows))
    package_col = _column(headers, "PACKAGE")
    package_name_col = _column(headers, "PACKAGE NAME")
    supplier_col = _column(headers, "SUPPLIER")
    main_col = _column(headers, "Main Description")
    amount_col = _column(headers, "AMOUNT (EUR)")
    status_col = _column(headers, "STATUS")
    type_col = _column(headers, "TYPE")
    description_col = _column(headers, "DESCRIPTION")
    vowd_col = _column(headers, "Cumulative VOWD")
    for raw in rows:
        row = list(raw)
        sid = _scope_id(_value(row, main_col), _value(row, package_col), _value(row, package_name_col))
        if sid not in CONTROL_SCOPES:
            continue
        amount = _number(_value(row, amount_col)) or 0.0
        vowd = _number(_value(row, vowd_col)) or 0.0
        status = _clean(_value(row, status_col)).lower()
        cost_type = _clean(_value(row, type_col)).lower()
        if status == "committed" and cost_type != "potential change":
            totals[sid]["committed"] += amount
        totals[sid]["vowd"] += vowd
        package = _clean(_value(row, package_col)) or "Unassigned"
        supplier = _clean(_value(row, supplier_col))
        key = (package, supplier)
        target = packages[sid].setdefault(key, {
            "package": package,
            "description": _clean(_value(row, package_name_col)) or _clean(_value(row, description_col)),
            "supplier": supplier,
            "forecast": 0.0,
            "vowd": 0.0,
            "potential_exposure": 0.0,
        })
        if status == "committed" and cost_type != "potential change":
            target["forecast"] += amount
        target["vowd"] += vowd
    output: Dict[str, list[Dict[str, Any]]] = {}
    for sid, grouped in packages.items():
        rows_out = list(grouped.values())
        for item in rows_out:
            item["potential_forecast"] = item["forecast"] + item["potential_exposure"]
        output[sid] = sorted(rows_out, key=lambda item: abs(item["forecast"]), reverse=True)
    return totals, output


def _exposure_data(workbook: Any) -> tuple[Dict[str, Dict[str, float]], Dict[str, list[Dict[str, Any]]]]:
    totals = {sid: {"pending": 0.0, "potential": 0.0} for sid in CONTROL_SCOPES}
    exposures = {sid: [] for sid in CONTROL_SCOPES}
    if "O&V Register" not in workbook.sheetnames:
        return totals, exposures
    sheet = workbook["O&V Register"]
    rows = sheet.iter_rows(values_only=True)
    headers = _headers(next(rows))
    package_col = _column(headers, "PACKAGE")
    package_name_col = _column(headers, "PACKAGE NAME")
    supplier_col = _column(headers, "SUPPLIER")
    main_col = _column(headers, "Main Description")
    amount_col = _column(headers, "AMOUNT (EUR)")
    type_col = _column(headers, "TYPE")
    status_col = _column(headers, "STATUS")
    description_col = _column(headers, "DESCRIPTION")
    criticality_col = _column(headers, "Criticality")
    item_col = _column(headers, "ITEM ID")
    cr_col = _column(headers, "CR Number")
    for raw in rows:
        row = list(raw)
        sid = _scope_id(_value(row, main_col), _value(row, package_col), _value(row, package_name_col))
        if sid not in CONTROL_SCOPES:
            continue
        cost_type = _clean(_value(row, type_col))
        amount = _number(_value(row, amount_col)) or 0.0
        if cost_type.lower() == "pending change":
            totals[sid]["pending"] += amount
        elif cost_type.lower() == "potential change":
            totals[sid]["potential"] += amount
        else:
            continue
        exposures[sid].append({
            "item_id": _clean(_value(row, item_col)),
            "package": _clean(_value(row, package_col)),
            "package_name": _clean(_value(row, package_name_col)),
            "description": _clean(_value(row, description_col)),
            "supplier": _clean(_value(row, supplier_col)),
            "status": _clean(_value(row, status_col)),
            "type": cost_type,
            "criticality": _clean(_value(row, criticality_col)),
            "cr_number": _clean(_value(row, cr_col)),
            "potential_exposure": amount,
        })
    for sid in CONTROL_SCOPES:
        exposures[sid].sort(key=lambda item: abs(item["potential_exposure"]), reverse=True)
    return totals, exposures


def normalize_cost_workbook(raw: bytes, source_name: str, previous: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    workbook = load_workbook(BytesIO(raw), data_only=True, read_only=True)
    report_date = _report_date(workbook)
    rail_rows, rail_packages = _budget_package_rows(workbook)
    tracker, packages = _tracker_data(workbook)
    exposure_totals, exposures = _exposure_data(workbook)
    previous_summary = (previous or {}).get("scope_summary") or {}

    summaries: Dict[str, Dict[str, Any]] = {}
    rail = rail_rows.get("rail") or {}
    rail_budget = _number(rail.get("budget"))
    rail_forecast = _number(rail.get("forecast"))
    rail_exposure = _number(rail.get("potential_exposure")) or exposure_totals["rail"]["potential"]
    rail_potential = _number(rail.get("potential_forecast"))
    if rail_potential is None and rail_forecast is not None:
        rail_potential = rail_forecast + rail_exposure
    summaries["rail"] = {
        "scope_id": "rail",
        "basis": "Budget per Package - Rails On Site Total",
        "budget": rail_budget,
        "change_order": (rail_forecast - rail_budget) if rail_forecast is not None and rail_budget is not None else None,
        "forecast": rail_forecast,
        "eac": rail_forecast,
        "potential_exposure": rail_exposure,
        "potential_forecast": rail_potential,
        "vowd": _number(rail.get("vowd")) if _number(rail.get("vowd")) is not None else tracker["rail"]["vowd"],
        "display_note": "Current Rail total from the dated Budget per Package control row.",
    }

    for sid, label in (("ponds", "Ponds and Culverts"), ("roads", "Roads and ditches")):
        budget = _number((previous_summary.get(sid) or {}).get("budget"))
        forecast = tracker[sid]["committed"] + exposure_totals[sid]["pending"]
        exposure = exposure_totals[sid]["potential"]
        summaries[sid] = {
            "scope_id": sid,
            "basis": f"Commitment tracker + O&V Register - {label}",
            "budget": budget,
            "change_order": (forecast - budget) if budget is not None else None,
            "forecast": forecast,
            "eac": forecast,
            "pending_changes": exposure_totals[sid]["pending"],
            "potential_exposure": exposure,
            "potential_forecast": forecast + exposure,
            "vowd": tracker[sid]["vowd"],
            "display_note": "Forecast uses committed value plus pending changes; potential exposure uses the active O&V Register. Budget is carried from the approved scope baseline because the current workbook does not publish a scope-level budget split.",
        }

    packages["rail"] = rail_packages or packages["rail"]
    for sid in CONTROL_SCOPES:
        exposure_by_package = defaultdict(float)
        for item in exposures[sid]:
            exposure_by_package[item.get("package") or "Unassigned"] += item["potential_exposure"]
        for item in packages[sid]:
            item["potential_exposure"] = exposure_by_package.get(item.get("package") or "Unassigned", item.get("potential_exposure") or 0.0)
            item["potential_forecast"] = (item.get("forecast") or 0.0) + item["potential_exposure"]

    return {
        "metadata": {
            "report_name": Path(source_name).stem,
            "source_file": source_name,
            "cutoff_month": report_date.strftime("%Y-%m"),
            "active_date": report_date.strftime("%Y-%m"),
            "data_date": report_date.strftime("%Y-%m-%d"),
            "currency": "EUR",
            "normalization_version": "2026-08-schema-aware-v1",
            "quality_note": "Rail is sourced from the dated package summary. Ponds and Roads use the current commitment and O&V registers; their approved scope budgets are carried forward because the redesigned workbook does not expose that split.",
        },
        "scope_summary": summaries,
        "packages_by_scope": packages,
        "components_by_scope": {sid: [] for sid in CONTROL_SCOPES},
        "exposures_by_scope": exposures,
    }
